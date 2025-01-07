from email.mime.image import MIMEImage
from toImage import toImage
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from concurrent.futures import ThreadPoolExecutor
from toPDF import toPDF

import os



path = "C:\\Users\\CoreCom\\Downloads\\final.xlsx"
work = openpyxl.load_workbook(path)
sheet = work.active


gmail_user = 'hjilani15@gmail.com'
gmail_password = 'sefs sffw mksu chca'
def send_email(i):
    gmail_receiver = sheet.cell(row=i, column=1).value
    print(gmail_receiver)
    if gmail_receiver is not None:
        name = sheet.cell(row=i, column=2).value
        msg = MIMEMultipart('alternative')
        msg['From'] = gmail_user
        msg['To'] = gmail_receiver
        msg['Subject'] = "Allotment Details for The Grand Debate'24"

        # Generate image
        output_image = toImage(name)
        
        if output_image:
            with open(output_image, 'rb') as file:
                img_data = file.read()
                img = MIMEImage(img_data, name=os.path.basename(output_image))
                msg.attach(img)
            
            try:
                server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
                server.login(gmail_user, gmail_password)
                server.sendmail(gmail_user, gmail_receiver, msg.as_string())
                server.close()
                sheet.cell(row=i, column=4).value = "Send"
                
                print(f"Email sent to {gmail_receiver}")
                work.save(path)
            except Exception as e:
                sheet.cell(row=i, column=4).value = 'Unsend'
                print(e)
                print(f"Did not send to {gmail_receiver}")
                work.save(path)
        else:
            print(f"Image generation failed for {name}")

  
with ThreadPoolExecutor(max_workers=3) as executor:
    futures = [executor.submit(send_email, i) for i in range(1, sheet.max_row + 1)]