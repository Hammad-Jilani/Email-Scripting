import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from concurrent.futures import ThreadPoolExecutor
import subprocess
from toPDF import toPDF
from certificate import certificates
from createCertificate import getCertificate


path = "C:\\Users\\CoreCom\\Downloads\\final.xlsx"

work = openpyxl.load_workbook(path)
sheet = work.active

gmail_user = ''
gmail_password = ''

def send_email(i):
  gmail_receiver = sheet.cell(row=i,column=1).value
  
  if gmail_receiver is not None:
    name = sheet.cell(row=i,column=2).value
    
    msg = MIMEMultipart('alternative')
    msg['From'] = gmail_user
    msg['To'] = gmail_receiver
    msg['Subject'] ="Coders Cup Participation Certificate"

    output = toPDF(name)
    # output = getCertificate(name)

    with open(output, 'rb') as file:
      attach = MIMEApplication(file.read(), _subtype='pdf')
      attach.add_header('Content-Disposition', 'attachment', filename=output)
      msg.attach(attach)
    
    try:
      server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
      server.login(gmail_user, gmail_password)
      server.sendmail(gmail_user, gmail_receiver, msg.as_string())
      server.close()
      sheet.cell(row=i,column=4).value="Send"
      
      print(f"Email send to {gmail_receiver}")
      work.save(path)

    except Exception as e:
      sheet.cell(row=i,column=4).value='Unsend'
      print(e)
      print(f"Did not send to {gmail_receiver}")
      work.save(path)
    
with ThreadPoolExecutor(max_workers=3) as executor:
    futures = [executor.submit(send_email, i) for i in range(1, sheet.max_row + 1)]
